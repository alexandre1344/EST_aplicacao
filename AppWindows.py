import sys
from datetime import datetime, timedelta
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QPushButton, QLabel, QLineEdit, 
                           QScrollArea, QMessageBox, QCheckBox, QFrame, QDialog,
                           QGridLayout, QSizePolicy)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QPalette, QColor, QIcon
from PyQt6.QtCore import Qt, QTimer
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pymongo import MongoClient
import os

# Estilos globais
STYLE = {
    'CORES': {
        'primaria': "#1E88E5",       # Azul principal
        'secundaria': "#FFA000",     # Laranja
        'sucesso': "#43A047",        # Verde
        'erro': "#E53935",          # Vermelho
        'fundo': "#121212",         # Preto
        'fundo_claro': "#1E1E1E",   # Cinza escuro
        'texto': "#FFFFFF",         # Branco
        'texto_secundario': "#B0BEC5" # Cinza claro
    },
    'FONTES': {
        'titulo': "font-size: 32px; font-weight: bold;",
        'subtitulo': "font-size: 24px; font-weight: bold;",
        'normal': "font-size: 14px;",
        'pequena': "font-size: 12px;"
    },
    'ESPACAMENTO': {
        'pequeno': 8,
        'medio': 16,
        'grande': 24
    },
    'BORDAS': {
        'raio': "8px"
    }
}

# Manter CORES como referência para compatibilidade
CORES = {
    'preto': STYLE['CORES']['fundo'],
    'preto_claro': STYLE['CORES']['fundo_claro'],
    'branco': STYLE['CORES']['texto'],
    'laranja': STYLE['CORES']['secundaria'],
    'vermelho': STYLE['CORES']['erro'],
    'verde': STYLE['CORES']['sucesso']
}

# Estilos de componentes
COMPONENT_STYLES = {
    'janela_principal': f"""
        QMainWindow {{
            background-color: {STYLE['CORES']['fundo']};
        }}
    """,
    'widget_container': f"""
        QWidget {{
            background-color: {STYLE['CORES']['fundo']};
        }}
    """,
    'botao_primario': f"""
        QPushButton {{
            background-color: {STYLE['CORES']['primaria']};
            color: {STYLE['CORES']['texto']};
            border: none;
            padding: 10px 20px;
            border-radius: {STYLE['BORDAS']['raio']};
            {STYLE['FONTES']['normal']}
        }}
        QPushButton:hover {{
            background-color: {STYLE['CORES']['primaria']}DD;
        }}
        QPushButton:pressed {{
            background-color: {STYLE['CORES']['primaria']}AA;
        }}
    """,
    'botao_secundario': f"""
        QPushButton {{
            background-color: {STYLE['CORES']['secundaria']};
            color: {STYLE['CORES']['texto']};
            border: none;
            padding: 10px 20px;
            border-radius: {STYLE['BORDAS']['raio']};
            {STYLE['FONTES']['normal']}
        }}
        QPushButton:hover {{
            background-color: {STYLE['CORES']['secundaria']}DD;
        }}
        QPushButton:pressed {{
            background-color: {STYLE['CORES']['secundaria']}AA;
        }}
    """,
    'input': f"""
        QLineEdit {{
            background-color: {STYLE['CORES']['fundo_claro']};
            color: {STYLE['CORES']['texto']};
            border: 1px solid {STYLE['CORES']['primaria']};
            padding: 10px;
            border-radius: {STYLE['BORDAS']['raio']};
            {STYLE['FONTES']['normal']}
        }}
        QLineEdit:focus {{
            border: 2px solid {STYLE['CORES']['primaria']};
        }}
    """,
    'label': f"""
        QLabel {{
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['normal']}
        }}
    """,
    'label_titulo': f"""
        QLabel {{
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['titulo']}
        }}
    """,
    'card': f"""
        QFrame {{
            background-color: {STYLE['CORES']['fundo_claro']};
            border-radius: {STYLE['BORDAS']['raio']};
            padding: {STYLE['ESPACAMENTO']['medio']}px;
        }}
    """
}

def get_download_folder():
    """Retorna o caminho da pasta Downloads no Windows"""
    return os.path.join(os.path.expanduser('~'), 'Downloads')

def get_database():
    """Conecta ao MongoDB e retorna a database"""
    try:
        connection_string = "mongodb+srv://dragonsky134:R9YaV9fRYkjEMczi@cluster0.arpj1.mongodb.net/fogo_e_lenha_db?retryWrites=true&w=majority"
        client = MongoClient(
            connection_string,
            serverSelectionTimeoutMS=5000,
            connectTimeoutMS=5000,
            socketTimeoutMS=5000,
            maxPoolSize=1
        )
        
        db = client.get_database()
        db.command('ping')
        
        return db
        
    except Exception as e:
        print(f"Erro detalhado na conexão MongoDB: {str(e)}")
        if "SSL: CERTIFICATE_VERIFY_FAILED" in str(e):
            raise Exception("Erro de certificado SSL. Verifique a data e hora do dispositivo.")
        elif "No servers found" in str(e):
            raise Exception("Não foi possível encontrar o servidor MongoDB. Verifique sua conexão.")
        else:
            raise Exception(f"Erro ao conectar ao banco de dados: {str(e)}")

# Primeiro, vamos adicionar uma classe base para configurações responsivas
class ResponsiveWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )

    def resizeEvent(self, event):
        # Calcula as proporções baseadas no tamanho da tela
        width = event.size().width()
        # Ajusta tamanhos mínimos baseados na largura da tela
        self.min_button_width = max(120, int(width * 0.15))
        self.min_input_width = max(200, int(width * 0.25))
        super().resizeEvent(event)

class ListaRegistrosDialog(QDialog):
    def __init__(self, parent=None, usuario=None):
        super().__init__(parent)
        self.usuario = usuario
        self.db = get_database()
        
        # Inicializar o timer
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.atualizar_pontos_tempo_real)
        self.timer.start(1000)
        
        self.init_ui()

    def init_ui(self):
        try:
            self.setWindowTitle("Lista de Registros")
            self.setMinimumWidth(600)
            
            layout = QVBoxLayout(self)
            layout.setSpacing(STYLE['ESPACAMENTO']['medio'])
            
            # Adiciona botão de atualização no topo
            btn_atualizar = QPushButton("↻ Atualizar Pontos")
            btn_atualizar.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['primaria']};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 10px 20px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {STYLE['CORES']['primaria']}DD;
                }}
            """)
            btn_atualizar.clicked.connect(self.atualizar_pontos_tempo_real)
            layout.addWidget(btn_atualizar)
            
            # Área de scroll para os registros
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            
            registros_widget = QWidget()
            self.lista_layout = QVBoxLayout(registros_widget)
            scroll.setWidget(registros_widget)
            
            layout.addWidget(scroll)
            
            # Botão de fechar
            btn_fechar = QPushButton("Fechar")
            btn_fechar.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['secundaria']};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 10px 20px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                }}
            """)
            btn_fechar.clicked.connect(self.close)
            
            layout.addWidget(btn_fechar, alignment=Qt.AlignmentFlag.AlignRight)
            
            self.carregar_registros()
            
        except Exception as e:
            print(f"Erro ao inicializar UI: {str(e)}")

    def carregar_registros(self):
        try:
            # Limpar lista atual
            while self.lista_layout.count():
                item = self.lista_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            
            colecao_anotacoes = self.db['anotacoes']
            
            # Buscar todas as anotações (removido o filtro de pago: False)
            registros = colecao_anotacoes.find().sort("data_criacao", -1)
            
            for registro in registros:
                item_widget = QFrame()
                item_widget.setStyleSheet(f"""
                    QFrame {{
                        background-color: {STYLE['CORES']['fundo']};
                        border-radius: {STYLE['BORDAS']['raio']};
                        padding: 8px;
                    }}
                """)
                
                item_layout = QHBoxLayout(item_widget)
                
                # Informações do registro
                info_layout = QVBoxLayout()
                
                # Adiciona a data junto com o usuário
                data = registro['data_criacao'].strftime("%d/%m/%Y")
                usuario_data_label = QLabel(f"Usuário: {registro.get('usuario', 'Desconhecido')} - Data: {data}")
                usuario_data_label.setStyleSheet(f"color: {STYLE['CORES']['texto_secundario']};")
                
                hora_placa = QLabel(f"{registro['hora']} - {registro['anotacao']}")
                hora_placa.setStyleSheet(f"color: {STYLE['CORES']['texto']};")
                
                # Adiciona o status (PAGO/PENDENTE)
                status = "PAGO" if registro.get('pago', False) else "PENDENTE"
                status_color = STYLE['CORES']['sucesso'] if registro.get('pago', False) else STYLE['CORES']['erro']
                status_label = QLabel(f"Status: {status}")
                status_label.setStyleSheet(f"color: {status_color}; font-weight: bold;")
                
                valor = QLabel(f"R$ {registro.get('pontos', 0):.2f}")
                valor.setStyleSheet(f"color: {STYLE['CORES']['sucesso']};")
                
                info_layout.addWidget(usuario_data_label)
                info_layout.addWidget(hora_placa)
                info_layout.addWidget(status_label)
                info_layout.addWidget(valor)
                
                item_layout.addLayout(info_layout)
                
                # Botões de ação (apenas se não estiver pago)
                if not registro.get('pago', False):
                    botoes_layout = QVBoxLayout()
                    
                    btn_pagar = QPushButton("✓ Marcar como Pago")
                    btn_pagar.setStyleSheet(f"""
                        QPushButton {{
                            background-color: {STYLE['CORES']['sucesso']};
                            color: {STYLE['CORES']['texto']};
                            border: none;
                            padding: 8px;
                            border-radius: {STYLE['BORDAS']['raio']};
                            font-weight: bold;
                        }}
                    """)
                    btn_pagar.clicked.connect(lambda checked, _id=registro['_id']: self.marcar_como_pago(_id))
                    botoes_layout.addWidget(btn_pagar)
                    
                    btn_excluir = QPushButton("× Excluir")
                    btn_excluir.setStyleSheet(f"""
                        QPushButton {{
                            background-color: {STYLE['CORES']['erro']};
                            color: {STYLE['CORES']['texto']};
                            border: none;
                            padding: 8px;
                            border-radius: {STYLE['BORDAS']['raio']};
                            font-weight: bold;
                        }}
                    """)
                    btn_excluir.clicked.connect(lambda checked, _id=registro['_id']: self.excluir_registro(_id))
                    botoes_layout.addWidget(btn_excluir)
                    
                    item_layout.addLayout(botoes_layout)
                
                self.lista_layout.addWidget(item_widget)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar registros: {str(e)}")

    def atualizar_pontos_tempo_real(self):
        """Atualiza os pontos das anotações pendentes"""
        try:
            colecao_anotacoes = self.db['anotacoes']
            
            # Buscar todas as anotações não pagas
            anotacoes = list(colecao_anotacoes.find({"pago": False}))
            
            alguma_atualizacao = False
            
            for anotacao in anotacoes:
                # Calcular novos pontos
                pontos_antigos = anotacao.get('pontos', 0)
                pontos_novos = self.parent().calcular_pontos(anotacao['data_criacao'])
                
                # Só atualiza se o valor mudou
                if abs(pontos_novos - pontos_antigos) >= 0.01:  # Diferença mínima de 1 centavo
                    colecao_anotacoes.update_one(
                        {"_id": anotacao['_id']},
                        {"$set": {"pontos": pontos_novos}}
                    )
                    alguma_atualizacao = True
            
            # Só recarrega a lista se houve alguma atualização
            if alguma_atualizacao:
                self.carregar_registros()
            
        except Exception as e:
            print(f"Erro ao atualizar pontos em tempo real: {str(e)}")
            import traceback
            traceback.print_exc()

    def marcar_como_pago(self, registro_id):
        try:
            colecao_anotacoes = self.db['anotacoes']
            resultado = colecao_anotacoes.update_one(
                {"_id": registro_id},
                {"$set": {"pago": True}}
            )
            
            if resultado.modified_count > 0:
                self.carregar_registros()
                if self.parent():
                    self.parent().carregar_anotacoes()
            else:
                QMessageBox.warning(self, "Erro", "Não foi possível marcar como pago")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao marcar como pago: {str(e)}")
    
    def excluir_registro(self, registro_id):
        try:
            resposta = QMessageBox.question(
                self,
                "Confirmar Exclusão",
                "Tem certeza que deseja excluir este registro?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if resposta == QMessageBox.StandardButton.Yes:
                colecao_anotacoes = self.db['anotacoes']
                resultado = colecao_anotacoes.delete_one({"_id": registro_id})
                
                if resultado.deleted_count > 0:
                    self.carregar_registros()
                    if self.parent():
                        self.parent().carregar_anotacoes()
                else:
                    QMessageBox.warning(self, "Erro", "Não foi possível excluir o registro")
                    
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao excluir registro: {str(e)}")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.usuario_atual = None
        self.setWindowTitle("Fogo e Lenha - Sistema de Gestão")
        self.setStyleSheet(f"background-color: {STYLE['CORES']['fundo']};")
        self.setMinimumSize(1024, 768)
        
        # Inicializar conexão com o banco de dados
        try:
            self.db = get_database()
        except Exception as e:
            print(f"Erro ao conectar ao banco de dados: {str(e)}")
            self.db = None
        
        # Widget central
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        # Layout principal
        self.main_layout = QVBoxLayout(self.central_widget)
        
        # Iniciar com a tela de login
        self.mostrar_login()

    def mostrar_login(self):
        # Limpar layout atual
        self.limpar_layout()
        
        # Criar e adicionar widget de login
        login_widget = LoginWidget(self)
        self.main_layout.addWidget(login_widget)

    def mostrar_anotacao(self, usuario):
        # Limpar layout atual
        self.limpar_layout()
        
        self.usuario_atual = usuario
        anotacao_widget = QWidget()
        layout = QVBoxLayout(anotacao_widget)
        layout.setSpacing(STYLE['ESPACAMENTO']['grande'])
        layout.setContentsMargins(STYLE['ESPACAMENTO']['grande'], 
                                STYLE['ESPACAMENTO']['grande'],
                                STYLE['ESPACAMENTO']['grande'], 
                                STYLE['ESPACAMENTO']['grande'])

        # Cabeçalho principal
        header = QFrame()
        header.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(STYLE['ESPACAMENTO']['grande'], 
                                       STYLE['ESPACAMENTO']['medio'],
                                       STYLE['ESPACAMENTO']['grande'], 
                                       STYLE['ESPACAMENTO']['medio'])

        # Informações do usuário
        info_container = QWidget()
        info_layout = QVBoxLayout(info_container)
        info_layout.setSpacing(4)
        
        label_bem_vindo = QLabel("Bem-vindo(a) ao Sistema")
        label_bem_vindo.setStyleSheet(f"""
            color: {STYLE['CORES']['texto_secundario']};
            {STYLE['FONTES']['normal']}
        """)
        
        nome_usuario = QLabel(usuario['nome'].upper())
        nome_usuario.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['titulo']}
            font-weight: bold;
        """)
        
        info_layout.addWidget(label_bem_vindo)
        info_layout.addWidget(nome_usuario)
        header_layout.addWidget(info_container, stretch=2)

        # Botões do header
        botoes_container = QWidget()
        botoes_layout = QHBoxLayout(botoes_container)
        botoes_layout.setSpacing(STYLE['ESPACAMENTO']['medio'])

        # Botão de pesquisa
        pesquisar_btn = QPushButton("🔍 Pesquisar Placas")
        pesquisar_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['primaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px 24px;
                border-radius: {STYLE['BORDAS']['raio']};
                {STYLE['FONTES']['normal']}
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {STYLE['CORES']['primaria']}DD;
            }}
        """)
        pesquisar_btn.clicked.connect(self.mostrar_pesquisa_placas)
        pesquisar_btn.setFixedWidth(150)

        voltar_btn = QPushButton("← Voltar")
        voltar_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px 24px;
                border-radius: {STYLE['BORDAS']['raio']};
                {STYLE['FONTES']['normal']}
                font-weight: bold;
            }}
        """)
        voltar_btn.clicked.connect(self.mostrar_login)
        voltar_btn.setFixedWidth(120)
        
        logout_btn = QPushButton("Sair ×")
        logout_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['erro']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px 24px;
                border-radius: {STYLE['BORDAS']['raio']};
                {STYLE['FONTES']['normal']}
                font-weight: bold;
            }}
        """)
        logout_btn.clicked.connect(self.mostrar_login)
        logout_btn.setFixedWidth(120)

        botoes_layout.addWidget(pesquisar_btn)
        botoes_layout.addWidget(voltar_btn)
        botoes_layout.addWidget(logout_btn)
        botoes_layout.setAlignment(Qt.AlignmentFlag.AlignRight)
        header_layout.addWidget(botoes_container, stretch=1)
        
        layout.addWidget(header)

        # Container principal dividido em duas colunas
        content_container = QWidget()
        content_layout = QHBoxLayout(content_container)
        content_layout.setSpacing(STYLE['ESPACAMENTO']['grande'])

        # Coluna da esquerda (Input)
        left_column = QWidget()
        left_layout = QVBoxLayout(left_column)
        left_layout.setSpacing(STYLE['ESPACAMENTO']['grande'])

        # Container para input
        input_container = QFrame()
        input_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        input_layout = QVBoxLayout(input_container)
        input_layout.setSpacing(STYLE['ESPACAMENTO']['medio'])
        
        input_label = QLabel(" Digite a placa do veículo")
        input_label.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['subtitulo']}
        """)
        input_layout.addWidget(input_label)

        self.input_anotacao = QLineEdit()
        self.input_anotacao.setStyleSheet(f"""
            QLineEdit {{
                background-color: {STYLE['CORES']['fundo']};
                color: {STYLE['CORES']['texto']};
                border: 2px solid {STYLE['CORES']['secundaria']};
                padding: 12px;
                border-radius: {STYLE['BORDAS']['raio']};
                {STYLE['FONTES']['subtitulo']}
            }}
        """)
        self.input_anotacao.setPlaceholderText("Ex: ABC1234")
        self.input_anotacao.setFixedHeight(50)
        input_layout.addWidget(self.input_anotacao)
        
        btn_adicionar = QPushButton("+ Adicionar Placa")
        btn_adicionar.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px 24px;
                border-radius: {STYLE['BORDAS']['raio']};
                {STYLE['FONTES']['normal']}
                font-weight: bold;
            }}
        """)
        btn_adicionar.clicked.connect(self.adicionar_anotacao)
        btn_adicionar.setFixedHeight(50)
        input_layout.addWidget(btn_adicionar)
        
        left_layout.addWidget(input_container)

        # Total na coluna esquerda
        total_container = QFrame()
        total_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['sucesso']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        total_layout = QHBoxLayout(total_container)
        
        total_label = QLabel("💰 Total do Dia:")
        total_label.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['subtitulo']}
            font-weight: bold;
        """)
        
        self.label_total = QLabel("R$ 0,00")
        self.label_total.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['titulo']}
            font-weight: bold;
        """)
        
        total_layout.addWidget(total_label)
        total_layout.addWidget(self.label_total, alignment=Qt.AlignmentFlag.AlignRight)
        
        left_layout.addWidget(total_container)
        left_layout.addStretch()

        # Coluna da direita (Lista)
        right_column = QFrame()
        right_column.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        right_layout = QVBoxLayout(right_column)
        
        btn_lista = QPushButton("📋 Ver Lista Completa")
        btn_lista.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 20px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-size: 18px;
                font-weight: bold;
            }}
        """)
        btn_lista.clicked.connect(lambda: self.mostrar_lista_registros(usuario))
        right_layout.addWidget(btn_lista)
        
        # Adicionar colunas ao container principal
        content_layout.addWidget(left_column, stretch=1)
        content_layout.addWidget(right_column, stretch=1)
        
        layout.addWidget(content_container)
        self.main_layout.addWidget(anotacao_widget)
        
        # Carregar anotações existentes
        self.carregar_anotacoes()

    def mostrar_lista_registros(self, usuario):
        dialog = ListaRegistrosDialog(self, usuario)
        dialog.exec()

    def mostrar_admin(self, usuario):
        # Limpar layout atual
        self.limpar_layout()
        
        self.usuario_atual = usuario
        admin_widget = AdminWidget(self)
        self.main_layout.addWidget(admin_widget)

    def limpar_layout(self):
        # Remover todos os widgets do layout principal
        while self.main_layout.count():
            item = self.main_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    def adicionar_anotacao(self):
        anotacao = self.input_anotacao.text().strip()
        if not anotacao:
            QMessageBox.warning(self, "Erro", "Digite uma placa!")
            return

        try:
            db = get_database()
            colecao_anotacoes = db['anotacoes']
            
            hora_criacao = datetime.now()
            pontos_inicial = self.calcular_pontos(hora_criacao)
            
            documento = {
                "anotacao": anotacao,
                "hora": hora_criacao.strftime("%H:%M"),
                "data_criacao": hora_criacao,
                "usuario": self.usuario_atual['nome'],
                "pontos": pontos_inicial,
                "pago": False
            }
            
            resultado = colecao_anotacoes.insert_one(documento)
            
            if resultado.inserted_id:
                self.input_anotacao.clear()
                self.carregar_anotacoes()
                QMessageBox.information(self, "Sucesso", "Placa adicionada com sucesso!")
            else:
                QMessageBox.warning(self, "Erro", "Falha ao salvar a placa")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao adicionar placa: {str(e)}")

    def carregar_anotacoes(self):
        try:
            db = get_database()
            colecao_anotacoes = db['anotacoes']
            
            hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            amanha = hoje + timedelta(days=1)
            
            anotacoes = colecao_anotacoes.find({
                "data_criacao": {
                    "$gte": hoje,
                    "$lt": amanha
                }
            }).sort("data_criacao", -1)
            
            total_pontos = 0
            
            for anotacao in anotacoes:
                # Atualiza os pontos antes de processar
                pontos_atualizados = self.calcular_pontos(anotacao['data_criacao'])
                
                # Atualiza o valor no banco de dados
                colecao_anotacoes.update_one(
                    {"_id": anotacao['_id']},
                    {"$set": {"pontos": pontos_atualizados}}
                )
                
                # Soma ao total se estiver pago
                if anotacao.get('pago', False):
                    total_pontos += pontos_atualizados
            
            # Atualiza o total na interface
            self.label_total.setText(f"R$ {total_pontos:.2f}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar anotações: {str(e)}")

    def adicionar_item_lista(self, anotacao):
        item_widget = QFrame()
        item_widget.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: 4px;
                padding: 8px;
                margin: 2px;
            }}
        """)
        
        item_layout = QHBoxLayout()
        
        hora = QLabel(anotacao['hora'])
        hora.setStyleSheet(f"color: {STYLE['CORES']['texto']};")
        hora.setFixedWidth(100)
        
        texto = QLabel(anotacao['anotacao'])
        texto.setStyleSheet(f"color: {STYLE['CORES']['texto']};")
        
        pontos = QLabel(f"R$ {anotacao.get('pontos', 0):.2f}")
        pontos.setStyleSheet(f"color: {STYLE['CORES']['sucesso']};")
        pontos.setFixedWidth(100)
        
        item_layout.addWidget(hora)
        item_layout.addWidget(texto)
        item_layout.addWidget(pontos)
        
        item_widget.setLayout(item_layout)
        self.lista_layout.addWidget(item_widget)
        self.anotacoes.append(anotacao)

    def mostrar_pesquisa_placas(self):
        """Abre o diálogo de pesquisa de placas"""
        dialog = PesquisaPlacasDialog(self, self.usuario_atual)
        dialog.exec()

    def calcular_pontos(self, hora_criacao):
        """Calcula os pontos baseado no tempo decorrido desde a criação"""
        try:
            # Primeiro, verifica se a anotação está paga
            colecao_anotacoes = self.db['anotacoes']
            anotacao = colecao_anotacoes.find_one({"data_criacao": hora_criacao})
            
            # Se a anotação estiver paga, retorna os pontos atuais sem modificar
            if anotacao and anotacao.get('pago', False):
                return anotacao.get('pontos', 5)
            
            # Se não estiver paga, calcula normalmente
            hora_atual = datetime.now()
            
            # Calcula a diferença total em horas entre a criação e agora
            diferenca = hora_atual - hora_criacao
            diferenca_horas = diferenca.total_seconds() / 3600
            
            # Cálculo progressivo: começa com 5 e soma 2 para cada hora que passou
            pontos_total = 5  # valor inicial
            for hora in range(int(diferenca_horas)):
                pontos_total += 2
            
            return pontos_total
            
        except Exception as e:
            print(f"Erro ao calcular pontos: {str(e)}")
            return 5  # Retorna o valor mínimo em caso de erro

class AdminWidget(ResponsiveWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.nivel_atual = "usuario"
        self.init_ui()

    def init_ui(self):
        # Layout principal
        layout = QVBoxLayout()
        layout.setSpacing(STYLE['ESPACAMENTO']['grande'])
        layout.setContentsMargins(20, 20, 20, 20)

        # Container principal
        main_container = QFrame()
        main_container.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )
        main_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        main_layout = QVBoxLayout(main_container)
        main_layout.setSpacing(STYLE['ESPACAMENTO']['grande'])

        # Cabeçalho
        header = QFrame()
        header.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['medio']}px;
            }}
        """)
        header_layout = QHBoxLayout(header)
        
        nome_usuario = QLabel(f"Administrador: {self.main_window.usuario_atual['nome']}")
        nome_usuario.setStyleSheet(f"""
            color: {STYLE['CORES']['secundaria']};
            font-size: 24px;
            font-weight: bold;
        """)
        header_layout.addWidget(nome_usuario)

        logout_btn = QPushButton("Sair")
        logout_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['erro']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-weight: bold;
                min-width: 120px;
            }}
        """)
        logout_btn.clicked.connect(self.logout)
        header_layout.addWidget(logout_btn)
        main_layout.addWidget(header)

        # Container de Cadastro
        cadastro_container = QFrame()
        cadastro_container.setSizePolicy(
            QSizePolicy.Policy.Preferred,
            QSizePolicy.Policy.Preferred
        )
        cadastro_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        cadastro_layout = QVBoxLayout(cadastro_container)
        cadastro_layout.setSpacing(STYLE['ESPACAMENTO']['medio'])

        # Título do Cadastro
        titulo_cadastro = QLabel("Cadastro de Novo Usuário")
        titulo_cadastro.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            font-size: 20px;
            font-weight: bold;
            padding-bottom: 10px;
            border-bottom: 2px solid {STYLE['CORES']['secundaria']};
        """)
        cadastro_layout.addWidget(titulo_cadastro, alignment=Qt.AlignmentFlag.AlignHCenter)

        # Form container
        form_container = QFrame()
        form_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: 20px;
                margin: 10px 50px;  /* Margem horizontal para centralizar */
            }}
        """)
        form_layout = QVBoxLayout(form_container)
        form_layout.setSpacing(20)  # Aumentar espaçamento entre os campos

        # Campos de entrada
        self.usuario_input = QLineEdit()
        self.senha_input = QLineEdit()
        
        for label_text, input_widget in [
            ("Usuário:", self.usuario_input),
            ("Senha:", self.senha_input)
        ]:
            container = QVBoxLayout()
            container.setSpacing(10)
            
            label = QLabel(label_text)
            label.setStyleSheet(f"""
                color: {STYLE['CORES']['texto']};
                font-size: 16px;
                font-weight: bold;
                margin-left: 5px;  /* Alinhar com o campo de texto */
            """)
            
            input_widget.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {STYLE['CORES']['fundo']};
                    color: {STYLE['CORES']['texto']};
                    border: 2px solid {STYLE['CORES']['secundaria']};
                    padding: 12px 15px;  /* Padding horizontal aumentado */
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-size: 16px;
                    min-width: 300px;  /* Largura mínima definida */
                }}
                QLineEdit:focus {{
                    border: 2px solid {STYLE['CORES']['primaria']};
                    background-color: {STYLE['CORES']['fundo_claro']};
                }}
            """)
            input_widget.setFixedHeight(45)
            input_widget.setMinimumWidth(400)  # Largura mínima maior
            
            if label_text == "Usuário:":
                input_widget.setPlaceholderText("Digite o nome do usuário")
            else:
                input_widget.setPlaceholderText("Digite a senha")
                input_widget.setEchoMode(QLineEdit.EchoMode.Password)
            
            container.addWidget(label)
            container.addWidget(input_widget)
            form_layout.addLayout(container)

        cadastro_layout.addWidget(form_container)

        # Container para botões
        botoes_container = QVBoxLayout()
        botoes_container.setSpacing(STYLE['ESPACAMENTO']['medio'])

        # Primeira linha de botões
        primeira_linha = QHBoxLayout()
        primeira_linha.setSpacing(STYLE['ESPACAMENTO']['medio'])

        # Botões de cadastro
        for btn_text, btn_color, btn_slot in [
            ("👤 Nível: Usuário", STYLE['CORES']['secundaria'], self.mudar_nivel_acesso),
            ("✓ Cadastrar Usuário", STYLE['CORES']['sucesso'], self.cadastrar_usuario)
        ]:
            btn = QPushButton(btn_text)
            if btn_text.startswith("👤"):
                self.nivel_btn = btn
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {btn_color};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 12px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                    font-size: 16px;
                    min-width: 200px;
                }}
            """)
            btn.clicked.connect(btn_slot)
            primeira_linha.addWidget(btn)

        botoes_container.addLayout(primeira_linha)

        # Segunda linha com botão de gerenciar
        gerenciar_btn = QPushButton("👥 Gerenciar Usuários")
        gerenciar_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-weight: bold;
                font-size: 16px;
            }}
        """)
        gerenciar_btn.clicked.connect(self.mostrar_gerenciar_usuarios)
        botoes_container.addWidget(gerenciar_btn)

        # Separador
        separador = QFrame()
        separador.setFrameShape(QFrame.Shape.HLine)
        separador.setStyleSheet(f"""
            background-color: {STYLE['CORES']['secundaria']};
            margin: 10px 0px;
        """)
        botoes_container.addWidget(separador)

        # Botões de relatório
        botoes_relatorio = QHBoxLayout()
        botoes_relatorio.setSpacing(STYLE['ESPACAMENTO']['medio'])

        for btn_text, btn_color, btn_slot in [
            ("📊 Relatório de Vendas", STYLE['CORES']['secundaria'], self.gerar_relatorio),
            ("📑 Relatório Excel", STYLE['CORES']['sucesso'], self.gerar_relatorio_excel)
        ]:
            btn = QPushButton(btn_text)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {btn_color};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 12px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                    font-size: 16px;
                    min-width: 200px;
                }}
            """)
            btn.clicked.connect(btn_slot)
            botoes_relatorio.addWidget(btn)

        botoes_container.addLayout(botoes_relatorio)
        cadastro_layout.addLayout(botoes_container)

        main_layout.addWidget(cadastro_container)
        layout.addWidget(main_container)
        self.setLayout(layout)

    def cadastrar_usuario(self):
        usuario = self.usuario_input.text().strip()
        senha = self.senha_input.text().strip()

        if not usuario or not senha:
            QMessageBox.warning(self, "Erro", "Usuário e senha são obrigatórios!")
            return

        try:
            db = get_database()
            colecao_usuarios = db['usuarios']
            
            # Verificar se usuário já existe
            if colecao_usuarios.find_one({"nome": usuario}):
                QMessageBox.warning(self, "Erro", "Usuário já existe!")
                return
            
            # Criar novo usuário
            novo_usuario = {
                "nome": usuario,
                "senha": senha,
                "nivel_acesso": self.nivel_atual,
                "data_criacao": datetime.now()
            }
            
            resultado = colecao_usuarios.insert_one(novo_usuario)
            
            if resultado.inserted_id:
                QMessageBox.information(self, "Sucesso", "Usuário cadastrado com sucesso!")
                self.usuario_input.clear()
                self.senha_input.clear()
            else:
                QMessageBox.warning(self, "Erro", "Falha ao cadastrar usuário")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao cadastrar usuário: {str(e)}")

    def mudar_nivel_acesso(self):
        if self.nivel_atual == "usuario":
            self.nivel_atual = "administrador"
            self.nivel_btn.setText("👤 Nível: Administrador")
            self.nivel_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['erro']};
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    font-size: 14px;
                    font-weight: bold;
                    border-radius: 4px;
                }}
            """)
        else:
            self.nivel_atual = "usuario"
            self.nivel_btn.setText("👤 Nível: Usuário")
            self.nivel_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['secundaria']};
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    font-size: 14px;
                    font-weight: bold;
                    border-radius: 4px;
                }}
            """)

    def gerar_relatorio(self):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("Selecionar Registros para Relatório")
            dialog.setMinimumWidth(600)
            
            layout = QVBoxLayout()
            
            # Área de scroll para os registros
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            
            registros_widget = QWidget()
            self.registros_layout = QVBoxLayout()
            registros_widget.setLayout(self.registros_layout)
            
            try:
                db = get_database()
                colecao_anotacoes = db['anotacoes']
                
                # Modificado o pipeline para incluir apenas registros pagos
                pipeline = [
                    {
                        "$match": {
                            "pago": True  # Filtra apenas registros pagos
                        }
                    },
                    {
                        "$group": {
                            "_id": {
                                "$dateToString": {
                                    "format": "%d/%m/%Y",
                                    "date": "$data_criacao"
                                }
                            },
                            "total_placas": {"$sum": 1},
                            "total_ganho": {"$sum": "$pontos"},
                            "registros": {"$push": "$$ROOT"}
                        }
                    },
                    {"$sort": {"_id": -1}}
                ]
                
                resultados = list(colecao_anotacoes.aggregate(pipeline))
                self.dados_registros = {}
                
                if not resultados:
                    # Se não houver registros pagos
                    info_label = QLabel("Não há registros pagos para gerar relatório")
                    info_label.setStyleSheet(f"""
                        color: {STYLE['CORES']['texto_secundario']};
                        {STYLE['FONTES']['normal']}
                        padding: 20px;
                    """)
                    info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.registros_layout.addWidget(info_label)
                else:
                    for grupo in resultados:
                        container = QWidget()
                        container_layout = QHBoxLayout()
                        
                        checkbox = QCheckBox()
                        info = QLabel(
                            f"{grupo['_id']} - {grupo['total_placas']} placas pagas - "
                            f"R$ {grupo['total_ganho']:.2f}"
                        )
                        info.setStyleSheet(f"color: {STYLE['CORES']['texto']};")
                        
                        container_layout.addWidget(checkbox)
                        container_layout.addWidget(info)
                        container.setLayout(container_layout)
                        
                        self.registros_layout.addWidget(container)
                        self.dados_registros[grupo['_id']] = {
                            'checkbox': checkbox,
                            'dados': grupo
                        }
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao carregar registros: {str(e)}")
                return
            
            scroll.setWidget(registros_widget)
            layout.addWidget(scroll)
            
            # Botões
            btn_layout = QHBoxLayout()
            
            selecionar_todos_btn = QPushButton("Selecionar Todos")
            selecionar_todos_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['secundaria']};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 10px 20px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {STYLE['CORES']['secundaria']}DD;
                }}
            """)
            
            gerar_btn = QPushButton("Gerar Relatório")
            gerar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['sucesso']};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 10px 20px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {STYLE['CORES']['sucesso']}DD;
                }}
            """)
            
            cancelar_btn = QPushButton("Cancelar")
            cancelar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {STYLE['CORES']['erro']};
                    color: {STYLE['CORES']['texto']};
                    border: none;
                    padding: 10px 20px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {STYLE['CORES']['erro']}DD;
                }}
            """)
            
            selecionar_todos_btn.clicked.connect(self.selecionar_todos_registros)
            gerar_btn.clicked.connect(lambda: self.salvar_relatorio_mongodb(dialog))
            cancelar_btn.clicked.connect(dialog.reject)
            
            btn_layout.addWidget(selecionar_todos_btn)
            btn_layout.addWidget(gerar_btn)
            btn_layout.addWidget(cancelar_btn)
            
            layout.addLayout(btn_layout)
            dialog.setLayout(layout)
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao preparar seleção: {str(e)}")

    def gerar_relatorio_excel(self):
        try:
            db = get_database()
            if db is None:
                QMessageBox.critical(self, "Erro", "Não foi possível conectar ao banco de dados")
                return
            
            colecao_relatorios = db['relatorios']
            relatorio = colecao_relatorios.find_one(sort=[("data_geracao", -1)])
            
            if relatorio is None:
                QMessageBox.warning(self, "Erro", "Nenhum relatório encontrado")
                return
            
            # Criar workbook Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Ganhos"
            
            # Estilos
            header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            total_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Título do relatório
            ws['A1'] = "RELATÓRIO DE GANHOS"
            ws.merge_cells('A1:C1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Data de geração
            data_geracao = relatorio['data_geracao'].strftime("%d/%m/%Y %H:%M:%S")
            ws['A2'] = f"Data de geração: {data_geracao}"
            ws.merge_cells('A2:C2')
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Linha em branco
            ws.append([])
            
            # Cabeçalhos
            headers = ['Horário', 'Placas', 'Ganho (R$)']
            ws.append(headers)
            
            # Formatar cabeçalhos
            for col in range(1, 4):
                cell = ws.cell(row=4, column=col)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Adicionar dados
            row_num = 5
            total_ganhos = 0
            
            for registro in relatorio['registros']:
                # Adicionar data como subseção
                ws.append([f"Data: {registro['_id']}"])
                ws.merge_cells(f'A{row_num}:C{row_num}')
                ws.cell(row=row_num, column=1).font = Font(bold=True)
                row_num += 1
                
                # Adicionar registros do dia
                for detalhe in registro['registros']:
                    hora = detalhe['hora']
                    placa = detalhe['anotacao']
                    valor = float(detalhe['pontos'])
                    
                    ws.append([hora, placa, valor])
                    
                    # Formatar células
                    for col in range(1, 4):
                        cell = ws.cell(row=row_num, column=col)
                        cell.border = border
                        if col == 3:  # Coluna de valor
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                        elif col == 1:  # Coluna de hora
                            cell.alignment = Alignment(horizontal='center')
                    
                    total_ganhos += valor
                    row_num += 1
                
                # Linha em branco entre dias
                ws.append([])
                row_num += 1

            # Total geral
            ws.append(['TOTAL GERAL', '', total_ganhos])
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = total_fill
                cell.border = border
                cell.font = Font(bold=True)
                if col == 3:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = Alignment(horizontal='right')

            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15

            # Salvar arquivo
            nome_arquivo = f"relatorio_ganhos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caminho_completo = os.path.join(get_download_folder(), nome_arquivo)
            
            wb.save(caminho_completo)
            QMessageBox.information(self, "Sucesso", f"Relatório Excel gerado com sucesso!\nSalvo em: {caminho_completo}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar relatório Excel: {str(e)}")

    def salvar_relatorio_mongodb(self, dialog):
        try:
            registros_selecionados = [
                dados['dados'] for data, dados in self.dados_registros.items()
                if dados['checkbox'].isChecked()
            ]
            
            if not registros_selecionados:
                QMessageBox.warning(self, "Erro", "Selecione pelo menos um registro")
                return

            relatorio = {
                "data_geracao": datetime.now(),
                "gerado_por": self.main_window.usuario_atual['nome'],
                "registros": registros_selecionados
            }

            db = get_database()
            colecao_relatorios = db['relatorios']
            resultado = colecao_relatorios.insert_one(relatorio)

            if resultado.inserted_id:
                dialog.accept()
                QMessageBox.information(
                    self, 
                    "Sucesso", 
                    f"Relatório gerado com sucesso!\nID: {resultado.inserted_id}"
                )
            else:
                QMessageBox.warning(self, "Erro", "Falha ao salvar o relatório")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar relatório: {str(e)}")

    def selecionar_todos_registros(self):
        algum_selecionado = any(dados['checkbox'].isChecked() for dados in self.dados_registros.values())
        for dados in self.dados_registros.values():
            dados['checkbox'].setChecked(not algum_selecionado)

    def mostrar_gerenciar_usuarios(self):
        dialog = GerenciarUsuariosDialog(self)
        dialog.exec()

    def logout(self):
        if hasattr(self.main_window, 'logout'):
            self.main_window.logout()
        else:
            self.main_window.mostrar_login()

# Primeiro, vamos criar a nova classe para o popup de gerenciamento
class GerenciarUsuariosDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = get_database()
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Gerenciar Usuários")
        self.setMinimumWidth(int(QApplication.primaryScreen().size().width() * 0.4))        
        self.setMinimumHeight(int(QApplication.primaryScreen().size().height() * 0.6))
        
        layout = QVBoxLayout(self)
        layout.setSpacing(STYLE['ESPACAMENTO']['medio'])
        
        # Container principal
        main_container = QFrame()
        main_container.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )
        main_layout = QVBoxLayout(main_container)
        
        # Título
        titulo = QLabel("👥 Gerenciamento de Usuários")
        titulo.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            font-size: 24px;
            font-weight: bold;
            padding-bottom: 10px;
            border-bottom: 2px solid {STYLE['CORES']['secundaria']};
        """)
        main_layout.addWidget(titulo)
        
        # Área de scroll para os usuários
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
        """)
        
        self.lista_widget = QWidget()
        self.lista_layout = QVBoxLayout(self.lista_widget)
        scroll.setWidget(self.lista_widget)
        main_layout.addWidget(scroll)
        
        layout.addWidget(main_container)
        
        # Botão de fechar
        btn_fechar = QPushButton("Fechar")
        btn_fechar.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 10px 20px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-weight: bold;
            }}
        """)
        btn_fechar.clicked.connect(self.close)
        
        layout.addWidget(btn_fechar, alignment=Qt.AlignmentFlag.AlignRight)
        
        self.carregar_usuarios()
        
    def carregar_usuarios(self):
        try:
            while self.lista_layout.count():
                item = self.lista_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
                    
            colecao_usuarios = self.db['usuarios']
            usuarios = colecao_usuarios.find().sort("nome", 1)
            
            for usuario in usuarios:
                item_widget = QFrame()
                item_widget.setStyleSheet(f"""
                    QFrame {{
                        background-color: {STYLE['CORES']['fundo']};
                        border-radius: {STYLE['BORDAS']['raio']};
                        padding: 12px;
                        margin: 4px;
                    }}
                """)
                
                item_layout = QHBoxLayout(item_widget)
                
                info_layout = QVBoxLayout()
                nome = QLabel(f"Nome: {usuario['nome']}")
                nivel = QLabel(f"Nível: {usuario.get('nivel_acesso', 'usuário')}")
                for label in [nome, nivel]:
                    label.setStyleSheet(f"color: {STYLE['CORES']['texto']};")
                    info_layout.addWidget(label)
                
                item_layout.addLayout(info_layout)
                
                # Botões de ação
                btn_excluir = QPushButton("Excluir")
                btn_excluir.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {STYLE['CORES']['erro']};
                        color: {STYLE['CORES']['texto']};
                        border: none;
                        padding: 8px;
                        border-radius: {STYLE['BORDAS']['raio']};
                        font-weight: bold;
                    }}
                """)
                btn_excluir.clicked.connect(lambda checked, u=usuario: self.excluir_usuario(u))
                item_layout.addWidget(btn_excluir)
                
                self.lista_layout.addWidget(item_widget)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar usuários: {str(e)}")
            
    def excluir_usuario(self, usuario):
        try:
            resposta = QMessageBox.question(
                self,
                "Confirmar Exclusão",
                f"Tem certeza que deseja excluir o usuário {usuario['nome']}?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if resposta == QMessageBox.StandardButton.Yes:
                colecao_usuarios = self.db['usuarios']
                resultado = colecao_usuarios.delete_one({"_id": usuario['_id']})
                
                if resultado.deleted_count > 0:
                    self.carregar_usuarios()
                else:
                    QMessageBox.warning(self, "Erro", "Não foi possível excluir o usuário")
                    
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao excluir usuário: {str(e)}")

class LoginWidget(ResponsiveWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(STYLE['ESPACAMENTO']['grande'])
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Container principal
        main_container = QFrame()
        main_container.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )
        main_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['grande']}px;
            }}
        """)
        main_layout = QVBoxLayout(main_container)
        main_layout.setSpacing(STYLE['ESPACAMENTO']['grande'])
        
        # Título
        titulo = QLabel("🚗 Sistema de Estacionamento")
        titulo.setStyleSheet(f"""
            color: {STYLE['CORES']['texto']};
            {STYLE['FONTES']['titulo']}
            padding-bottom: {STYLE['ESPACAMENTO']['pequeno']}px;
            border-bottom: 2px solid {STYLE['CORES']['secundaria']};
        """)
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(titulo)
        
        # Form container
        form_container = QFrame()
        form_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: 20px;
                margin: 10px 50px;
            }}
        """)
        form_layout = QVBoxLayout(form_container)
        form_layout.setSpacing(20)
        
        # Campos de entrada
        self.usuario_input = QLineEdit()
        self.senha_input = QLineEdit()
        
        for label_text, input_widget in [
            ("Usuário:", self.usuario_input),
            ("Senha:", self.senha_input)
        ]:
            container = QVBoxLayout()
            container.setSpacing(10)
            
            label = QLabel(label_text)
            label.setStyleSheet(f"""
                color: {STYLE['CORES']['texto']};
                font-size: 16px;
                font-weight: bold;
                margin-left: 5px;
            """)
            
            input_widget.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {STYLE['CORES']['fundo_claro']};
                    color: {STYLE['CORES']['texto']};
                    border: 2px solid {STYLE['CORES']['secundaria']};
                    padding: 12px 15px;
                    border-radius: {STYLE['BORDAS']['raio']};
                    font-size: 16px;
                    min-width: 300px;
                }}
                QLineEdit:focus {{
                    border: 2px solid {STYLE['CORES']['primaria']};
                }}
            """)
            input_widget.setFixedHeight(45)
            input_widget.setMinimumWidth(400)
            
            if label_text == "Usuário:":
                input_widget.setPlaceholderText("Digite seu usuário")
            else:
                input_widget.setPlaceholderText("Digite sua senha")
                input_widget.setEchoMode(QLineEdit.EchoMode.Password)
            
            container.addWidget(label)
            container.addWidget(input_widget)
            form_layout.addLayout(container)
        
        # Botão de login
        btn_login = QPushButton("Entrar")
        btn_login.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 12px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-weight: bold;
                font-size: 16px;
                min-width: 200px;
            }}
        """)
        btn_login.clicked.connect(self.fazer_login)
        form_layout.addWidget(btn_login, alignment=Qt.AlignmentFlag.AlignCenter)
        
        main_layout.addWidget(form_container)
        layout.addWidget(main_container)
        self.setLayout(layout)
        
    def fazer_login(self):
        usuario = self.usuario_input.text().strip()
        senha = self.senha_input.text().strip()
        
        if not usuario or not senha:
            QMessageBox.warning(self, "Erro", "Usuário e senha são obrigatórios!")
            return
        
        try:
            db = get_database()
            colecao_usuarios = db['usuarios']
            
            usuario_encontrado = colecao_usuarios.find_one({
                "nome": usuario,
                "senha": senha
            })
            
            if usuario_encontrado:
                if usuario_encontrado.get('nivel_acesso') == 'administrador':
                    self.main_window.mostrar_admin(usuario_encontrado)
                else:
                    self.main_window.mostrar_anotacao(usuario_encontrado)
            else:
                QMessageBox.warning(self, "Erro", "Usuário ou senha inválidos!")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao fazer login: {str(e)}")

def get_responsive_style(width):
    return {
        'FONTES': {
            'titulo': f'font-size: {max(24, int(width * 0.02))}px;',
            'subtitulo': f'font-size: {max(18, int(width * 0.015))}px;',
            'texto': f'font-size: {max(14, int(width * 0.01))}px;'
        },
        'ESPACAMENTO': {
            'pequeno': max(8, int(width * 0.01)),
            'medio': max(16, int(width * 0.015)),
            'grande': max(24, int(width * 0.02))
        }
    }

# Nova classe para o diálogo de pesquisa
class PesquisaPlacasDialog(QDialog):
    def __init__(self, parent=None, usuario=None):
        super().__init__(parent)
        self.usuario = usuario
        self.db = get_database()
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Pesquisar Placas")
        self.setMinimumWidth(600)
        self.setMinimumHeight(400)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(STYLE['ESPACAMENTO']['medio'])
        
        # Campo de pesquisa
        search_container = QFrame()
        search_container.setStyleSheet(f"""
            QFrame {{
                background-color: {STYLE['CORES']['fundo_claro']};
                border-radius: {STYLE['BORDAS']['raio']};
                padding: {STYLE['ESPACAMENTO']['medio']}px;
            }}
        """)
        search_layout = QHBoxLayout(search_container)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Digite a placa para pesquisar...")
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {STYLE['CORES']['fundo']};
                color: {STYLE['CORES']['texto']};
                border: 2px solid {STYLE['CORES']['primaria']};
                padding: 10px;
                border-radius: {STYLE['BORDAS']['raio']};
                {STYLE['FONTES']['normal']}
            }}
        """)
        
        btn_pesquisar = QPushButton("🔍 Pesquisar")
        btn_pesquisar.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['primaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 10px 20px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-weight: bold;
            }}
        """)
        btn_pesquisar.clicked.connect(self.pesquisar)
        
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(btn_pesquisar)
        
        layout.addWidget(search_container)
        
        # Área de resultados
        self.resultados_widget = QWidget()
        self.resultados_layout = QVBoxLayout(self.resultados_widget)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.resultados_widget)
        scroll.setStyleSheet(f"""
            QScrollArea {{
                border: none;
                background-color: transparent;
            }}
        """)
        
        layout.addWidget(scroll)
        
        # Botão de fechar
        btn_fechar = QPushButton("Fechar")
        btn_fechar.setStyleSheet(f"""
            QPushButton {{
                background-color: {STYLE['CORES']['secundaria']};
                color: {STYLE['CORES']['texto']};
                border: none;
                padding: 10px 20px;
                border-radius: {STYLE['BORDAS']['raio']};
                font-weight: bold;
            }}
        """)
        btn_fechar.clicked.connect(self.close)
        
        layout.addWidget(btn_fechar, alignment=Qt.AlignmentFlag.AlignRight)
        
        # Conectar evento de Enter no campo de pesquisa
        self.search_input.returnPressed.connect(self.pesquisar)

    def pesquisar(self):
        # Limpar resultados anteriores
        while self.resultados_layout.count():
            item = self.resultados_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
                
        placa = self.search_input.text().strip()
        if not placa:
            return
            
        try:
            colecao_anotacoes = self.db['anotacoes']
            resultados = colecao_anotacoes.find({
                "anotacao": {"$regex": placa, "$options": "i"}
            }).sort("data_criacao", -1)
            
            for registro in resultados:
                item = QFrame()
                item.setStyleSheet(f"""
                    QFrame {{
                        background-color: {STYLE['CORES']['fundo_claro']};
                        border-radius: {STYLE['BORDAS']['raio']};
                        padding: {STYLE['ESPACAMENTO']['medio']}px;
                        margin: 2px;
                    }}
                """)
                
                item_layout = QVBoxLayout(item)
                
                # Informações do registro
                data = registro['data_criacao'].strftime("%d/%m/%Y")
                hora = registro['hora']
                placa = registro['anotacao']
                usuario = registro.get('usuario', 'Desconhecido')
                status = "PAGO" if registro.get('pago', False) else "PENDENTE"
                valor = registro.get('pontos', 0)
                
                info_label = QLabel(
                    f"Data: {data} | Hora: {hora}\n"
                    f"Placa: {placa}\n"
                    f"Usuário: {usuario}\n"
                    f"Status: {status} | Valor: R$ {valor:.2f}"
                )
                info_label.setStyleSheet(f"color: {STYLE['CORES']['texto']};")
                
                item_layout.addWidget(info_label)
                self.resultados_layout.addWidget(item)
            
            if self.resultados_layout.count() == 0:
                no_results = QLabel("Nenhum resultado encontrado")
                no_results.setStyleSheet(f"""
                    color: {STYLE['CORES']['texto_secundario']};
                    {STYLE['FONTES']['normal']}
                    padding: 20px;
                """)
                no_results.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.resultados_layout.addWidget(no_results)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao pesquisar: {str(e)}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())